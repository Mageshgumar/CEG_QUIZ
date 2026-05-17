#!/usr/bin/env python3
"""Teacher dashboard for managing tests and viewing student performance."""

from functools import wraps
import io
import json
import os

from flask import Flask, render_template, request, redirect, url_for, session, flash, jsonify, send_file
from werkzeug.security import generate_password_hash, check_password_hash

from openpyxl import Workbook

from config import TEACHER_USERNAME, TEACHER_PASSWORD, TEACHERS_FILE
from quiz import (
    load_tests,
    load_all_tests,
    save_tests,
    add_test,
    parse_test_file,
    set_active_test,
    set_test_inactive,
    set_all_tests_inactive,
    get_test_by_id,
    update_test,
    delete_test,
    build_test,
)
from user_data import UserDataManager

app = Flask(__name__)
app.secret_key = os.environ.get("DASHBOARD_SECRET", "cegquiz_dashboard_secret")
API_KEY = os.environ.get("API_KEY", "")

user_manager = UserDataManager()


def _load_teachers() -> list[dict]:
    if os.path.exists(TEACHERS_FILE):
        try:
            with open(TEACHERS_FILE) as fp:
                data = json.load(fp)
            return data if isinstance(data, list) else []
        except json.JSONDecodeError:
            return []
    return []


def _save_teachers(teachers: list[dict]) -> None:
    with open(TEACHERS_FILE, "w") as fp:
        json.dump(teachers, fp, indent=2)


def _find_teacher(username: str, teachers: list[dict]) -> dict | None:
    lookup = username.strip().lower()
    for teacher in teachers:
        if str(teacher.get("username", "")).lower() == lookup:
            return teacher
    return None


def _authenticate_teacher(username: str, password: str) -> bool:
    if username == TEACHER_USERNAME and password == TEACHER_PASSWORD:
        return True

    teachers = _load_teachers()
    if teachers:
        teacher = _find_teacher(username, teachers)
        return bool(teacher and check_password_hash(teacher.get("password_hash", ""), password))
    return False


def _register_teacher(username: str, password: str) -> tuple[bool, str]:
    teachers = _load_teachers()
    if _find_teacher(username, teachers):
        return False, "Username already registered."

    teachers.append(
        {
            "username": username.strip(),
            "password_hash": generate_password_hash(password),
        }
    )
    _save_teachers(teachers)
    return True, "Registration successful. Please log in."


def _current_teacher_username() -> str:
    return str(session.get("teacher_username") or TEACHER_USERNAME).strip()


def login_required(view_func):
    @wraps(view_func)
    def wrapper(*args, **kwargs):
        if not session.get("teacher_logged_in"):
            return redirect(url_for("login"))
        return view_func(*args, **kwargs)

    return wrapper


def _api_key_required(view_func):
    @wraps(view_func)
    def wrapper(*args, **kwargs):
        if not API_KEY:
            return jsonify({"error": "API key not configured"}), 403
        header_key = request.headers.get("X-API-Key", "")
        if header_key != API_KEY:
            return jsonify({"error": "Unauthorized"}), 403
        return view_func(*args, **kwargs)

    return wrapper


def _attempts_for_test(test_id: str):
    attempts = user_manager.load_attempts(_current_teacher_username())
    return [a for a in attempts if a.get("test_id") == test_id]


def _write_row(sheet, row_index: int, values: list[str | int]) -> None:
    for col_index, value in enumerate(values, start=1):
        sheet.cell(row=row_index, column=col_index, value=value)


def _excel_response(workbook: Workbook, filename: str):
    stream = io.BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return send_file(
        stream,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/api/tests", methods=["GET", "POST"])
@_api_key_required
def api_tests():
    if request.method == "GET":
        return jsonify(load_all_tests())

    data = request.get_json(silent=True)
    if not isinstance(data, list):
        return jsonify({"error": "Expected list of tests"}), 400
    save_tests(data, TEACHER_USERNAME)
    return jsonify({"ok": True})


@app.route("/api/attempts", methods=["GET", "POST"])
@_api_key_required
def api_attempts():
    if request.method == "GET":
        return jsonify(user_manager.load_all_attempts())

    data = request.get_json(silent=True)
    if not isinstance(data, dict):
        return jsonify({"error": "Expected attempt object"}), 400
    user_manager.save_attempt(data, teacher_username=data.get("teacher_username"))
    return jsonify({"ok": True})


@app.route("/api/attempts/<attempt_id>", methods=["DELETE"])
@_api_key_required
def api_attempts_delete(attempt_id: str):
    deleted, removed = user_manager.delete_attempt_by_id(attempt_id, TEACHER_USERNAME)
    return jsonify({"deleted": deleted, "removed": removed})


@app.route("/api/attempts/by-test/<test_id>", methods=["DELETE"])
@_api_key_required
def api_attempts_delete_by_test(test_id: str):
    removed_count = user_manager.delete_attempts_by_test_id(test_id, TEACHER_USERNAME)
    return jsonify({"removed_count": removed_count})


@app.route("/api/parents/<username>", methods=["GET"])
@_api_key_required
def api_parent_get(username: str):
    chat_id = user_manager.get_parent_chat_id(username, TEACHER_USERNAME)
    return jsonify({"chat_id": chat_id})


@app.route("/api/parents", methods=["POST"])
@_api_key_required
def api_parent_set():
    data = request.get_json(silent=True)
    if not isinstance(data, dict):
        return jsonify({"error": "Expected parent payload"}), 400
    username = str(data.get("username", "")).strip()
    chat_id = data.get("chat_id")
    teacher_username = str(data.get("teacher_username", TEACHER_USERNAME)).strip() or TEACHER_USERNAME
    if not username or chat_id is None:
        return jsonify({"error": "username and chat_id required"}), 400
    user_manager.register_parent(username, int(chat_id), teacher_username)
    return jsonify({"ok": True})


@app.route("/")
def index():
    if session.get("teacher_logged_in"):
        return redirect(url_for("dashboard"))
    return redirect(url_for("login"))


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")

        if _authenticate_teacher(username, password):
            session["teacher_logged_in"] = True
            session["teacher_username"] = username
            return redirect(url_for("dashboard"))

        flash("Invalid username or password.", "error")

    return render_template("login.html")


@app.route("/register", methods=["GET", "POST"])
def register():
    if session.get("teacher_logged_in"):
        return redirect(url_for("dashboard"))

    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")
        confirm_password = request.form.get("confirm_password", "")

        if len(username) < 3:
            flash("Username must be at least 3 characters.", "error")
            return render_template("register.html")

        if len(password) < 6:
            flash("Password must be at least 6 characters.", "error")
            return render_template("register.html")

        if password != confirm_password:
            flash("Passwords do not match.", "error")
            return render_template("register.html")

        ok, message = _register_teacher(username, password)
        if ok:
            flash(message, "success")
            return redirect(url_for("login"))

        flash(message, "error")

    return render_template("register.html")


@app.route("/logout")
@login_required
def logout():
    session.clear()
    return redirect(url_for("login"))


@app.route("/dashboard")
@login_required
def dashboard():
    teacher_username = _current_teacher_username()
    tests = load_tests(teacher_username)
    attempts = user_manager.load_attempts(teacher_username)

    attendance_by_test = {}
    for attempt in attempts:
        tid = attempt.get("test_id")
        attendance_by_test.setdefault(tid, 0)
        attendance_by_test[tid] += 1

    return render_template(
        "dashboard.html",
        tests=tests,
        attendance_by_test=attendance_by_test,
    )


@app.route("/test/<test_id>")
@login_required
def test_details(test_id: str):
    tests = load_tests(_current_teacher_username())
    test = next((t for t in tests if t.get("id") == test_id), None)
    if test is None:
        flash("Test not found.", "error")
        return redirect(url_for("dashboard"))

    attempts = _attempts_for_test(test_id)

    search_query = request.args.get("q", "").strip()
    sort_by = request.args.get("sort_by", "submitted_at")
    order = request.args.get("order", "desc")

    if search_query:
        needle = search_query.lower()
        attempts = [
            a for a in attempts
            if needle in str(((a.get("student") or {}).get("name", "")).lower())
            or needle in str(((a.get("student") or {}).get("roll", "")).lower())
        ]

    valid_sort = {"name", "roll", "submitted_at", "score"}
    if sort_by not in valid_sort:
        sort_by = "submitted_at"
    reverse = order != "asc"

    def sort_key(attempt: dict):
        student = attempt.get("student") or {}
        if sort_by == "name":
            return str(student.get("name", "")).lower()
        if sort_by == "roll":
            return str(student.get("roll", "")).lower()
        if sort_by == "score":
            return int(attempt.get("score", 0) or 0)
        return str(attempt.get("submitted_at", ""))

    attempts.sort(key=sort_key, reverse=reverse)

    selected_sort_submitted = "selected" if sort_by == "submitted_at" else ""
    selected_sort_name = "selected" if sort_by == "name" else ""
    selected_sort_roll = "selected" if sort_by == "roll" else ""
    selected_sort_score = "selected" if sort_by == "score" else ""
    selected_order_asc = "selected" if order == "asc" else ""
    selected_order_desc = "selected" if order == "desc" else ""

    return render_template(
        "test_details.html",
        test=test,
        attempts=attempts,
        search_query=search_query,
        sort_by=sort_by,
        order=order,
        selected_sort_submitted=selected_sort_submitted,
        selected_sort_name=selected_sort_name,
        selected_sort_roll=selected_sort_roll,
        selected_sort_score=selected_sort_score,
        selected_order_asc=selected_order_asc,
        selected_order_desc=selected_order_desc,
    )


@app.route("/download-test-report/<test_id>")
@login_required
def download_test_report(test_id: str):
    test = get_test_by_id(test_id, _current_teacher_username())
    if test is None:
        flash("Test not found.", "error")
        return redirect(url_for("dashboard"))

    attempts = _attempts_for_test(test_id)
    attempts.sort(key=lambda a: str(a.get("submitted_at", "")), reverse=True)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Test Report"

    headers = ["Student Name", "Roll no", "Phone no", "Score", "Submitted At"]
    _write_row(sheet, 1, headers)

    for idx, attempt in enumerate(attempts, start=2):
        student = attempt.get("student") or {}
        total_marks = attempt.get("total_marks", attempt.get("total_questions", 0))
        score = f"{attempt.get('score', 0)}/{total_marks}"
        _write_row(
            sheet,
            idx,
            [
                student.get("name", ""),
                student.get("roll", ""),
                student.get("phone", ""),
                score,
                attempt.get("submitted_at", ""),
            ],
        )

    safe_name = (test.get("name") or "test").strip().replace(" ", "_")
    filename = f"{safe_name}_students_report.xlsx"
    return _excel_response(workbook, filename)


@app.route("/attempt/<attempt_id>")
@login_required
def attempt_details(attempt_id: str):
    attempts = user_manager.load_attempts(_current_teacher_username())
    attempt = next((a for a in attempts if a.get("attempt_id") == attempt_id), None)
    if attempt is None:
        flash("Student attempt not found.", "error")
        return redirect(url_for("dashboard"))

    total = attempt.get("total_marks", attempt.get("total_questions", 0))
    score = attempt.get("score", 0)
    percentage = (score / total * 100) if total else 0

    return render_template(
        "attempt_details.html",
        attempt=attempt,
        percentage=percentage,
    )


@app.route("/download-attempt-report/<attempt_id>")
@login_required
def download_attempt_report(attempt_id: str):
    attempts = user_manager.load_attempts(_current_teacher_username())
    attempt = next((a for a in attempts if a.get("attempt_id") == attempt_id), None)
    if attempt is None:
        flash("Student attempt not found.", "error")
        return redirect(url_for("dashboard"))

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Attempt Report"

    headers = ["Question no", "Question", "Student Answer", "Correct Answer", "Status"]
    _write_row(sheet, 1, headers)

    answers = attempt.get("answers") or []
    for idx, ans in enumerate(answers, start=2):
        status = "Correct" if ans.get("is_correct") else "Wrong"
        _write_row(
            sheet,
            idx,
            [
                idx - 1,
                ans.get("question", ""),
                ans.get("user_answer", ""),
                ans.get("correct_answer", ""),
                status,
            ],
        )

    student = attempt.get("student") or {}
    safe_name = (student.get("name") or "student").strip().replace(" ", "_")
    filename = f"{safe_name}_attempt_report.xlsx"
    return _excel_response(workbook, filename)


@app.route("/delete-attempt/<attempt_id>", methods=["POST"])
@login_required
def delete_attempt(attempt_id: str):
    deleted, removed = user_manager.delete_attempt_by_id(attempt_id, _current_teacher_username())
    if not deleted:
        flash("Attempt not found.", "error")
        return redirect(url_for("dashboard"))

    flash("Student attempt deleted successfully.", "success")
    test_id = str((removed or {}).get("test_id", "")).strip()
    if test_id:
        return redirect(url_for("test_details", test_id=test_id))
    return redirect(url_for("dashboard"))


@app.route("/create-test", methods=["GET", "POST"])
@login_required
def create_test():
    if request.method == "POST":
        test_name = request.form.get("test_name", "").strip()
        test_file = request.files.get("test_file")
        make_active = request.form.get("make_active") == "on"
        one_time = request.form.get("one_time") == "on"

        if not test_file or not test_file.filename:
            flash("Please upload a test text file.", "error")
            return render_template("create_test.html")

        if not test_name:
            flash("Please enter a test name.", "error")
            return render_template("create_test.html")

        try:
            timer_seconds = int(request.form.get("timer_seconds", "30") or "30")
            random_count = int(request.form.get("random_count", "0") or "0")
            mark_correct = int(request.form.get("mark_correct", "1") or "1")
            mark_incorrect = int(request.form.get("mark_incorrect", "0") or "0")
            content = test_file.read().decode("utf-8")
            questions = parse_test_file(content)
            teacher_username = _current_teacher_username()
            test = build_test(
                name=test_name,
                questions=questions,
                timer_seconds=timer_seconds,
                random_count=random_count,
                one_time=one_time,
                mark_correct=mark_correct,
                mark_incorrect=mark_incorrect,
                is_active=make_active,
                teacher_username=teacher_username,
            )
            add_test(test, make_active=make_active, teacher_username=teacher_username)
            flash("Test created successfully.", "success")
            return redirect(url_for("dashboard"))
        except Exception as exc:
            flash(f"Failed to create test: {exc}", "error")

    return render_template("create_test.html")


@app.route("/set-active/<test_id>", methods=["POST"])
@login_required
def set_active(test_id: str):
    if set_active_test(test_id, _current_teacher_username()):
        flash("Test marked active.", "success")
    else:
        flash("Test not found.", "error")
    return redirect(url_for("dashboard"))


@app.route("/set-inactive/<test_id>", methods=["POST"])
@login_required
def set_inactive(test_id: str):
    if set_test_inactive(test_id, _current_teacher_username()):
        flash("Test marked inactive.", "success")
    else:
        flash("Test not found.", "error")
    return redirect(url_for("dashboard"))


@app.route("/set-all-inactive", methods=["POST"])
@login_required
def set_all_inactive():
    if set_all_tests_inactive(_current_teacher_username()):
        flash("All tests are now inactive.", "success")
    else:
        flash("No tests found.", "error")
    return redirect(url_for("dashboard"))


@app.route("/edit-test/<test_id>", methods=["GET", "POST"])
@login_required
def edit_test(test_id: str):
    teacher_username = _current_teacher_username()
    test = get_test_by_id(test_id, teacher_username)
    if test is None:
        flash("Test not found.", "error")
        return redirect(url_for("dashboard"))

    if request.method == "POST":
        test_name = request.form.get("test_name", "").strip()
        make_active = request.form.get("make_active") == "on"
        one_time = request.form.get("one_time") == "on"
        test_file = request.files.get("test_file")

        if not test_name:
            flash("Please enter a test name.", "error")
            return render_template("edit_test.html", test=test)

        if not test_file or not test_file.filename:
            flash("Please upload the updated test file.", "error")
            return render_template("edit_test.html", test=test)

        try:
            timer_seconds = int(request.form.get("timer_seconds", str(test.get("timer_seconds", 30))) or "30")
            random_count = int(request.form.get("random_count", str(test.get("random_count", 0))) or "0")
            mark_correct = int(request.form.get("mark_correct", str(test.get("mark_correct", 1))) or "1")
            mark_incorrect = int(request.form.get("mark_incorrect", str(test.get("mark_incorrect", 0))) or "0")
            content = test_file.read().decode("utf-8")
            questions = parse_test_file(content)
            updated = update_test(
                test_id=test_id,
                name=test_name,
                timer_seconds=timer_seconds,
                random_count=random_count,
                questions=questions,
                one_time=one_time,
                mark_correct=mark_correct,
                mark_incorrect=mark_incorrect,
                make_active=make_active,
                teacher_username=teacher_username,
            )
            if not updated:
                flash("Failed to update test.", "error")
                return redirect(url_for("dashboard"))

            removed_count = user_manager.delete_attempts_by_test_id(test_id, teacher_username)
            if removed_count:
                flash(
                    f"Test updated. Cleared {removed_count} previous attempts for this test.",
                    "success",
                )
            else:
                flash("Test updated. Ongoing student sessions for this test will be reset.", "success")
            return redirect(url_for("dashboard"))
        except Exception as exc:
            flash(f"Failed to edit test: {exc}", "error")

    return render_template("edit_test.html", test=test)


@app.route("/delete-test/<test_id>", methods=["POST"])
@login_required
def remove_test(test_id: str):
    if delete_test(test_id, _current_teacher_username()):
        removed_count = user_manager.delete_attempts_by_test_id(test_id, _current_teacher_username())
        if removed_count:
            flash(f"Test deleted. Removed {removed_count} related attempts.", "success")
        else:
            flash("Test deleted. Ongoing student sessions for this test will be reset.", "success")
    else:
        flash("Test not found.", "error")
    return redirect(url_for("dashboard"))


if __name__ == "__main__":
    port = int(os.environ.get("PORT", "10000"))
    app.run(host="0.0.0.0", port=port, debug=False)
